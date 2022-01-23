#!/usr/bin/python3

# Se importan los módulos. Se han añadido excepciones para que se instalen directamente con pip los módulos openpyxl, tqdm, requests y bs4 en caso de no estar instalados.

import sys
import subprocess
import os
import csv
import re
import ipaddress
from time import sleep

try:
    import requests
except:
    print('\nEl módulo requests no está instalado. Vamos a instalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    import requests

try:
    from bs4 import BeautifulSoup
except:
    print('\nEl módulo bs4 no está instalado. Vamos a instalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'bs4'])
    from bs4 import BeautifulSoup

try:
    from tqdm import tqdm
except:
    print('\nEl módulo tqdm no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'tqdm'])
    from tqdm import tqdm

try:
    import openpyxl
except:
    print('\nEl módulo openpyxl no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, NamedStyle, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

logo = '''
      ___           ___         ___           ___           ___           ___           
     /\__\         /\  \       /\  \         /\  \         /\__\         /\__\          
    /:/ _/_       /::\  \      \:\  \       /::\  \       /:/ _/_       /:/ _/_         
   /:/ /\  \     /:/\:\__\      \:\  \     /:/\:\__\     /:/ /\  \     /:/ /\__\        
  /:/ /::\  \   /:/ /:/  /  ___  \:\  \   /:/ /:/  /    /:/ /::\  \   /:/ /:/ _/_       
 /:/_/:/\:\__\ /:/_/:/  /  /\  \  \:\__\ /:/_/:/__/___ /:/__\/\:\__\ /:/_/:/ /\__\      
 \:\/:/ /:/  / \:\/:/  /   \:\  \ /:/  / \:\/:::::/  / \:\  \ /:/  / \:\/:/ /:/  /      
  \::/ /:/  /   \::/__/     \:\  /:/  /   \::/~~/~~~~   \:\  /:/  /   \::/_/:/  /       
   \/_/:/  /     \:\  \      \:\/:/  /     \:\~~\        \:\/:/  /     \:\/:/  /        
     /:/  /       \:\__\      \::/  /       \:\__\        \::/  /       \::/  /           
     \/__/         \/__/       \/__/         \/__/         \/__/         \/__/          

                                                                                                                                   
'''

print(logo)

file = input('\nIntroduce la ruta del archivo de texto en el que se encuentran las IPs que deseas comprobar (si estás en una máquina Windows doblar las barras invertidas para evitar error. e.g: C:\\\\Users\\\\usuario\\\\archivo.txt): ')

# Leer ips de archivo

ip_list = []
with open(file) as f:
    file_item = f.read()
    regex = r'(?:(?:2(?:[0-4][0-9]|5[0-5])|[0-1]?[0-9]?[0-9])\.){3}(?:(?:2([0-4][0-9]|5[0-5])|[0-1]?[0-9]?[0-9]))'
    matches = re.finditer(regex, file_item, re.MULTILINE)

    [ip_list.append(match.group())
     for matchNum, match in enumerate(matches, start=1)]

# Comprobar si hay IPs privadas en la lista y de ser así, sacarlas de ella

for i in ip_list:
    if ipaddress.ip_address(i).is_private is True:
        ip_list.remove(i)
        print('\n' + i + ' is a private IP address, and is only used in internal network environments. Any abusive activity you see coming from an internal IP is either coming from within your network itself, or is the result of an error or misconfiguration. ')

# Crear primera columna del csv, punto y coma como delimitador para que no den errores las comas de las respuestas

with open('resultados.csv', 'a') as file:
        file.write("IP;Resultado\n")
        file.close()

# Comprobar IPs en spur.us, y meter la salida en archivo

for i in ip_list:
    # Insertamos la IP en la primera columna del csv con punto y coma que luego vamos a usar como delimitador al convertir a excel
    with open('resultados.csv', 'a') as file:
        file.write(str(i)+";")
        file.close()
    url = 'https://spur.us/context/'+(i)
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    }
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, "html.parser")
    with open('respuesta.txt', 'a') as file:
        file.write(str(soup.prettify()))
        file.close()

    # Desbrozar respuesta
    with open('respuesta.txt', 'r') as f1, open('resultados.csv', 'a') as f2:
        for line in f1.readlines():
            if 'name="description"' in line:
                f2.write(line)
    f1.close()
    f2.close()

    with open('resultados.csv', 'rt') as file:
        x = file.read()
    with open('resultados.csv', 'wt') as file:
        x = x.replace('<meta content=\"', '')
        file.write(x)
    with open('resultados.csv', 'rt') as file:
        x = file.read()
    with open('resultados.csv', 'wt') as file:
        x = x.replace('\" name=\"description\">', '')
        file.write(x)

    os.remove('respuesta.txt')

    # Barra de progreso y sleep
    print('\nComprobando ' + str(i))
    for i in tqdm(range(100)):
        sleep(0.2)

# Convertir csv a excel, punto y coma como delimitador para que no nos den errores las comas de las respuestas

csv_data = []
with open('resultados.csv') as file_obj:
    reader = csv.reader(file_obj, delimiter=';')
    for row in reader:
        csv_data.append(row)
workbook = openpyxl.Workbook()
sheet = workbook.active
for row in csv_data:
    sheet.append(row)

# Ajustar tamaño de las columnas

MIN_WIDTH = 10
for i, column_cells in enumerate(sheet.columns, start=1):
    width = (
        length
        if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                          for cell in column_cells)) >= MIN_WIDTH
        else MIN_WIDTH
    )
    sheet.column_dimensions[get_column_letter(i)].width = width

# Primera fila en negrita

header = NamedStyle(name="header")
header.font = Font(bold=True)
header_row = sheet[1]
for cell in header_row:
    cell.style = header

# Pintar en rojo IPs de servicios de anonimato y de amarillo las de proxies

red_text = Font(color="0E0E0D")
red_fill = PatternFill(bgColor="cc0000")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("is part of",A1)))']
sheet.conditional_formatting.add("A1:O1000", rule)

yellow_text = Font(color="0E0E0D")
yellow_fill = PatternFill(bgColor="ffe65d")
dxf = DifferentialStyle(fill=yellow_fill, font=yellow_text)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("Proxy to anonymize",A1)))']
sheet.conditional_formatting.add("A1:O1000", rule)

# Congelar primera fila

sheet.freeze_panes = "A2"

# Añadir filtros

sheet.auto_filter.ref = "A1:B1000"

workbook.save('resultados.xlsx')
workbook.close()
